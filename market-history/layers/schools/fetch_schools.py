#!/usr/bin/env python3
"""NJ SCHOOL RATINGS per town, from the DOE's statewide NJSLA assessment results.

    python3 layers/schools/fetch_schools.py            # newest published school year
    python3 layers/schools/fetch_schools.py --check     # download + compare, write nothing
    python3 layers/schools/fetch_schools.py --year 2425 # pin a school year

Rebuilds the ratings in layers/schools/school_ratings.csv. This layer used to be
HAND-MAINTAINED, so the DOE's autumn release only landed if somebody remembered. Now it
re-derives itself, and `--year auto` (the default) walks forward from the known year, so
next autumn's file is picked up with no code change.

WHAT IT REBUILDS, AND WHAT IT PRESERVES. Only the measures move: the three
`*_rating_1_to_10`, the three `*_composite_proficiency_pct`, the three
`*_valid_test_scores`, and `school_year`. The TOWN -> DISTRICT ASSIGNMENT is preserved
verbatim, because it is curation this script cannot rebuild: schools are assigned by
attendance boundary, not by zip, and somebody worked out that (say) Garwood's elementary
is Garwood Boro while its high school is Clark Township. A refresh must never guess at
that. Same instinct as fetch_income.py preserving hand-written notes.

METHOD. For each district the DOE publishes one `District Total` / `All Students` row per
grade-and-subject file. Proficiency is L4 + L5 (met + exceeded expectations). Those are
combined into one composite per level, WEIGHTED BY VALID TEST SCORES, so a district's big
grades count for more than its small ones. The rating is then a statewide DECILE of that
composite among all NJ districts reporting at that level: 1 = bottom 10%, 10 = top 10%.

  elementary  ELA03-05, MAT03-05      <- reproduces the supplied file EXACTLY
  middle      ELA06-08, MAT06-08      <- reproduces the supplied file EXACTLY
  high school ELA09, ALG01, GEO01, ALG02   <- does NOT. See below.

Science (SC05/SC08/SC11) is EXCLUDED by default — `--with-science` includes it. The
hand-built 2024-25 file is reproduced by the ELA+MATH set, so that is the default; adding
science would silently redefine the metric mid-series.

HIGH SCHOOL IS NOT REFRESHED BY DEFAULT — AND THAT IS DELIBERATE.
Run against the same 2024-25 source the supplied file cites, elementary and middle
reproduce it to the decile on all 78 rows. High school reproduces 11 of 78, always
reading LOWER. That is not drift; it means the supplied high-school figures were built
from a wider set of tests than the spring NJSLA files contain. The tell is the test
counts: Woodbridge stores 2,769 valid scores where every spring high-school file for that
district totals 1,430. Roughly double. The likeliest missing piece is NJGPA, the 11th-grade
graduation assessment, which DOE publishes separately — unconfirmed.

So `--rebuild-high-school` exists but is OFF. Overwriting a number we cannot reconcile
would silently redefine the metric mid-series and break every comparison against the
ratings already published on the pages. Elementary and middle refresh; high school is
preserved verbatim until somebody confirms the source. Settle NJGPA and this flag becomes
the default.

THE CAVEAT THAT MUST RIDE WITH THE NUMBER. This is a DISTRICT-level proxy attached to a
zip via the town label. One zip can span districts at different levels, a large township
(Woodbridge: Colonia / Avenel / Iselin) has one label here and several catchments in
reality, and two houses on the same street can feed different elementary schools. For an
actual house, verify the boundary. `assignment_note` in the file says so, and that note is
not boilerplate — it is preserved here, never regenerated.
"""
import argparse
import csv
import os
import statistics
import urllib.error
import urllib.request

try:
    import openpyxl
except ImportError:
    raise SystemExit("!! needs openpyxl (local prep only, same as fetch_appreciation.py):\n"
                     "   python3 -m pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "school_ratings.csv")
CACHE = os.path.join(HERE, "raw")            # gitignored transient xlsx cache

BASE = "https://www.nj.gov/education/assessment/results/reports/{yy}/"
SPRING = BASE + "spring/{code}%20NJSLA%20DATA%20{label}.xlsx"

LEVELS = {
    "elementary":  ["ELA03", "ELA04", "ELA05", "MAT03", "MAT04", "MAT05"],
    "middle":      ["ELA06", "ELA07", "ELA08", "MAT06", "MAT07", "MAT08"],
    "high_school": ["ELA09", "ALG01", "GEO01", "ALG02"],
}
SCIENCE = {"elementary": ["SC05"], "middle": ["SC08"], "high_school": ["SC11"]}

KNOWN_YY = 2425          # the release we know exists; auto probes forward from here

# DOE sheet layout (header on row index 2)
C_COUNTY, C_DISTRICT, C_SCHOOL, C_SUBGROUP, C_SUBTYPE = 1, 3, 5, 6, 7
C_VALID, C_L4, C_L5 = 10, 15, 16


def yy_label(yy):
    """2425 -> ('2425', '2024-25')"""
    start = 2000 + int(str(yy)[:2])
    return str(yy), f"{start}-{str(yy)[2:]}"


def next_yy(yy):
    a, b = int(str(yy)[:2]), int(str(yy)[2:])
    return int(f"{a + 1:02d}{b + 1:02d}")


def published(yy):
    url = BASE.format(yy=yy_label(yy)[0])
    try:
        req = urllib.request.Request(url, method="HEAD", headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as r:
            return r.status == 200
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, OSError):
        return False


def newest_yy():
    yy = KNOWN_YY
    while published(next_yy(yy)):
        yy = next_yy(yy)
        if yy > KNOWN_YY + 1010:        # runaway guard (~10 years)
            break
    return yy


def fetch(code, yy):
    """Download (and cache) one grade-subject workbook. None if the DOE has no such file."""
    tag, label = yy_label(yy)
    os.makedirs(CACHE, exist_ok=True)
    path = os.path.join(CACHE, f"{code}_{tag}.xlsx")
    if not os.path.exists(path):
        url = SPRING.format(yy=tag, code=code, label=label)
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        try:
            with urllib.request.urlopen(req, timeout=180) as r, open(path, "wb") as fh:
                fh.write(r.read())
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, OSError) as e:
            print(f"    {code}: not available ({e})")
            return None
    return path


def read_book(path):
    """-> {(county, district): (valid_scores, proficient_pct)} from the District Total rows.

    Keyed by COUNTY TOO, and that is not decoration: three district names are reused
    across NJ counties — Washington Township (Gloucester/Morris/Warren), Greenwich
    (Cumberland/Gloucester/Warren) and Franklin (Hunterdon/Warren). Keyed by name alone
    the last county in the file silently wins, which handed Long Valley (Washington Twp,
    MORRIS) another county's scores and moved its rating four deciles.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3 or not row or len(row) <= C_L5:
            continue
        if row[C_SCHOOL] != "District Total" or row[C_SUBTYPE] != "All Students":
            continue
        try:
            valid = int(row[C_VALID])
            prof = float(row[C_L4]) + float(row[C_L5])
        except (TypeError, ValueError):
            continue                    # '*' = suppressed small cell
        if valid <= 0:
            continue
        out[(str(row[C_COUNTY]).strip(), str(row[C_DISTRICT]).strip())] = (valid, prof)
    wb.close()
    return out


def composites(codes, yy):
    """-> {district: (weighted_proficiency_pct, total_valid_scores)} for one level."""
    acc = {}
    for code in codes:
        path = fetch(code, yy)
        if not path:
            continue
        book = read_book(path)
        print(f"    {code}: {len(book)} districts")
        for key, (valid, prof) in book.items():
            w, n = acc.setdefault(key, [0.0, 0])
            acc[key] = [w + prof * valid, n + valid]
    return {d: (round(w / n, 1), n) for d, (w, n) in acc.items() if n > 0}


def deciles(comp):
    """composite pct -> 1..10 statewide decile. 1 = bottom 10%, 10 = top 10%."""
    vals = sorted(v for v, _ in comp.values())
    if not vals:
        return {}
    cuts = [statistics.quantiles(vals, n=10)[i] for i in range(9)] if len(vals) >= 10 else []
    out = {}
    for key, (pct, _) in comp.items():
        d = 1
        for c in cuts:
            if pct > c:
                d += 1
        out[key] = min(10, d)
    return out


def town_counties():
    """town -> DOE-style county name, from zips.json (which stores it upper-case)."""
    import json
    root = os.path.join(HERE, os.pardir, os.pardir)
    out = {}

    def walk(o):
        if isinstance(o, dict):
            if "name" in o and "county" in o:
                out[o["name"]] = str(o["county"]).title()
            for v in o.values():
                walk(v)
        elif isinstance(o, list):
            for v in o:
                walk(v)

    walk(json.load(open(os.path.join(root, "zips.json"))))
    return out


def resolve(name, county, comp):
    """Curated district name + the town's county -> the right (county, district) key."""
    hits = [k for k in comp if k[1] == name]
    if not hits:
        return None
    if len(hits) == 1:
        return hits[0]
    same = [k for k in hits if k[0] == county]      # ambiguous name: county decides
    return same[0] if same else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--year", default="auto", help="DOE school year like 2425, or 'auto'")
    ap.add_argument("--check", action="store_true", help="compare only, write nothing")
    ap.add_argument("--with-science", action="store_true", help="fold SC05/SC08/SC11 in")
    ap.add_argument("--rebuild-high-school", action="store_true",
                    help="also rewrite high school — REDEFINES the metric, see the docstring")
    args = ap.parse_args()

    levels = dict(LEVELS)
    if not args.rebuild_high_school:
        levels.pop("high_school")
        print("  high school PRESERVED (unreconciled — see docstring; "
              "--rebuild-high-school to override)")

    yy = newest_yy() if args.year == "auto" else int(args.year)
    tag, label = yy_label(yy)
    print(f"DOE school year: {label}" + ("  (newest published)" if args.year == "auto" else ""))
    if yy == KNOWN_YY:
        print("  no newer release than the one already in the file — numbers should not move")

    counties = town_counties()
    comp, dec = {}, {}
    for level, codes in levels.items():
        if args.with_science:
            codes = codes + SCIENCE[level]
        print(f"  {level}:")
        comp[level] = composites(codes, yy)
        if not comp[level]:
            raise SystemExit(f"!! {level} produced no districts — refusing to overwrite.")
        dec[level] = deciles(comp[level])
        print(f"    -> {len(comp[level])} districts with a composite")

    with open(OUT, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
        cols = list(rows[0].keys())

    moved, unmatched, same = [], set(), 0
    for r in rows:
        for level in levels:
            names = [d.strip() for d in (r[f"{level}_districts"] or "").split(";") if d.strip()]
            county = counties.get(r["town"], "")
            keys = {n: resolve(n, county, comp[level]) for n in names}
            got = [(comp[level][k], dec[level][k]) for k in keys.values() if k]
            for n, k in keys.items():
                if not k:
                    unmatched.add((level, n))
            if not got:
                continue
            # a zip served by >1 district at one level: weight them by test volume
            tot = sum(c[1] for (c, _) in got)
            pct = round(sum(c[0] * c[1] for (c, _) in got) / tot, 1) if tot else ""
            rating = max(d for (_, d) in got) if len(got) == 1 else \
                round(sum(d * c[1] for (c, d) in got) / tot)
            old = r[f"{level}_rating_1_to_10"]
            if old not in ("", None) and int(old) != int(rating):
                moved.append((r["town"], r["zip_code"], level, old, rating))
            else:
                same += 1
            r[f"{level}_rating_1_to_10"] = str(int(rating))
            r[f"{level}_composite_proficiency_pct"] = f"{pct}"
            r[f"{level}_valid_test_scores"] = str(tot)
        r["school_year"] = label
        r["source_url"] = BASE.format(yy=tag)

    print(f"\n  {len(rows)} zip rows | {same} ratings unchanged | {len(moved)} moved")
    for town, z, level, old, new in moved[:15]:
        print(f"     {town:<20} {z}  {level:<12} {old} -> {new}")
    if len(moved) > 15:
        print(f"     ... and {len(moved) - 15} more")
    if unmatched:
        print(f"  !! {len(unmatched)} curated district names had no DOE row "
              f"(a rename breaks the join — check these):")
        for level, n in sorted(unmatched)[:10]:
            print(f"     {level:<12} {n}")

    if args.check:
        print("--check: nothing written")
        return

    tmp = OUT + ".tmp"
    with open(tmp, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cols, quoting=csv.QUOTE_ALL)
        w.writeheader()
        w.writerows(rows)
    os.replace(tmp, OUT)
    print(f"wrote {os.path.relpath(OUT)} — {len(rows)} zips, {label}")


if __name__ == "__main__":
    main()
